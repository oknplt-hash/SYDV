import json
import os
import re
import sqlite3
from datetime import datetime
from io import BytesIO

from flask import (
    Flask,
    abort,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

SOCIAL_SECURITY_OPTIONS = [
    "G0",
    "G1",
    "Emekli",
    "65 Maaşı",
    "Engelli Maaşı",
    "Bağkur",
    "SGK",
]

DISABILITY_OPTIONS = ["Yok", "Var"]

ASSISTANCE_TYPES = [
    "Gida Yardimi",
    "Nakit Yardim",
    "Egitim Yardimi",
    "Giysi Yardimi",
    "Saglik Yardimi",
    "Kira Yardimi",
    "Komur Yardimi",
    "Tibbi Cihaz Yardimi",
    "Ev Onarim Yardimi",
    "Ev Esyasi Yardimi",
    "Aile Yardimi",
    "Tek Seferlik Yardim",
    "Yol Yardimi",
    "Universite Ogrencilerine Yonelik Yardim",
    "Sartli Egitim Saglik Yardimi",
    "Yasli Ayligi",
    "Engelli Ayligi",
    "Dogalgaz",
    "Diger Merkezi Yardimlar",
    "65'lik Maasi",
    "Diger",
]

CENTRAL_ASSISTANCE_OPTIONS = [
    "Engelli Aylığı",
    "Yaşlı Aylığı",
    "Şartlı Eğitim Sağlık",
    "Doğalgaz Tüketim Desteği",
    "Elektrik Tüketim Desteği",
    "SHÇEK",
    "E.V.E.K",
]

AGENDA_ASSISTANCE_TYPES = list(ASSISTANCE_TYPES)


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key")
    app.config["DATABASE"] = os.path.join(app.instance_path, "registry.sqlite3")
    os.makedirs(app.instance_path, exist_ok=True)

    def get_db():
        if "db" not in g:
            g.db = sqlite3.connect(app.config["DATABASE"])
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON;")
        return g.db

    @app.teardown_appcontext
    def close_db(error=None):
        db = g.pop("db", None)
        if db is not None:
            db.close()

    def init_db():
        db = get_db()
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS persons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_no TEXT NOT NULL UNIQUE,
                full_name TEXT NOT NULL,
                national_id TEXT,
                spouse_name TEXT,
                birth_date TEXT,
                household_size INTEGER,
                children_count INTEGER,
                student_count INTEGER,
                phone TEXT,
                address TEXT,
                social_security TEXT,
                disability_status TEXT,
                disability_rate TEXT,
                central_assistance TEXT,
                household_description TEXT,
                household_income REAL,
                per_capita_income REAL,
                profile_photo BLOB,
                profile_photo_filename TEXT,
                profile_photo_mimetype TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS assistance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                assistance_type TEXT NOT NULL,
                assistance_date TEXT,
                assistance_amount REAL,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS household_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                image_data BLOB NOT NULL,
                filename TEXT,
                mimetype TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS agendas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                meeting_date TEXT,
                description TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS agenda_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                agenda_id INTEGER NOT NULL,
                person_id INTEGER NOT NULL,
                application_date TEXT,
                assistance_type TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (agenda_id) REFERENCES agendas(id) ON DELETE CASCADE,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE
            );
            """
        )

        def ensure_column(table, column, definition):
            existing = db.execute(f"PRAGMA table_info({table})").fetchall()
            if not any(row["name"] == column for row in existing):
                db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition};")

        ensure_column("persons", "household_income", "REAL")
        ensure_column("persons", "per_capita_income", "REAL")
        ensure_column("persons", "birth_date", "TEXT")
        db.commit()

    def fetch_person(person_id):
        db = get_db()
        person = db.execute(
            "SELECT * FROM persons WHERE id = ?", (person_id,)
        ).fetchone()
        if person is None:
            abort(404, description="Kayıt bulunamadı.")
        return person

    def fetch_agenda(agenda_id):
        db = get_db()
        agenda = db.execute(
            "SELECT * FROM agendas WHERE id = ?", (agenda_id,)
        ).fetchone()
        if agenda is None:
            abort(404, description="Gündem bulunamadı.")
        return agenda

    with app.app_context():
        init_db()

    @app.route("/")
    def index():
        db = get_db()
        query = """
            SELECT id, file_no, full_name, phone, social_security, created_at, updated_at
            FROM persons
            ORDER BY datetime(created_at) DESC
        """
        persons = db.execute(query).fetchall()
        return render_template("index.html", persons=persons)

    @app.route("/persons/search")
    def search_persons():
        file_no = request.args.get("file_no", "").strip()
        full_name = request.args.get("full_name", "").strip()

        db = get_db()
        base_query = """
            SELECT id, file_no, full_name, phone, social_security, created_at, updated_at
            FROM persons
        """

        filters = []
        params = []
        if file_no:
            filters.append("file_no LIKE ?")
            params.append(f"{file_no}%")
        if full_name:
            filters.append("full_name LIKE ?")
            params.append(f"%{full_name}%")

        if filters:
            base_query += " WHERE " + " AND ".join(filters)

        base_query += " ORDER BY datetime(created_at) DESC LIMIT 100"

        def format_dt(value):
            if not value:
                return ""
            try:
                dt = datetime.fromisoformat(value)
            except ValueError:
                return value
            return dt.strftime("%d.%m.%Y %H:%M")

        rows = db.execute(base_query, params).fetchall()
        persons = [
            {
                "id": row["id"],
                "file_no": row["file_no"],
                "full_name": row["full_name"],
                "phone": row["phone"] or "",
                "social_security": row["social_security"] or "",
                "created_at": format_dt(row["created_at"]),
                "updated_at": format_dt(row["updated_at"]),
                "edit_url": url_for("edit_person", person_id=row["id"]),
                "delete_url": url_for("delete_person", person_id=row["id"]),
            }
            for row in rows
        ]

        return jsonify({"persons": persons})

    @app.route("/agendas")
    def list_agendas():
        db = get_db()
        agendas = db.execute(
            """
            SELECT a.*, COUNT(ai.id) AS item_count
            FROM agendas a
            LEFT JOIN agenda_items ai ON ai.agenda_id = a.id
            GROUP BY a.id
            ORDER BY datetime(a.created_at) DESC
            """
        ).fetchall()
        return render_template("agendas.html", agendas=agendas)

    @app.route("/agenda/new", methods=["GET", "POST"])
    def create_agenda():
        if request.method == "POST":
            title = request.form.get("title", "").strip()
            meeting_date = request.form.get("meeting_date", "").strip() or None
            description = None

            if not title:
                flash("Gündem başlığı zorunludur.", "danger")
                return redirect(request.url)

            now = datetime.utcnow().isoformat()
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO agendas (title, meeting_date, description, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (title, meeting_date, description, now, now),
            )
            db.commit()

            flash("Gündem oluşturuldu.", "success")
            return redirect(url_for("edit_agenda", agenda_id=cursor.lastrowid))

        return render_template(
            "agenda_form.html",
            agenda=None,
            items=[],
            assistance_types=AGENDA_ASSISTANCE_TYPES,
        )

    @app.route("/agenda/<int:agenda_id>/edit")
    def edit_agenda(agenda_id):
        agenda = fetch_agenda(agenda_id)
        db = get_db()
        items = db.execute(
            """
            SELECT ai.id, ai.agenda_id, ai.application_date, ai.assistance_type,
                   ai.notes, ai.created_at, p.id AS person_id, p.file_no, p.full_name,
                   p.phone, p.address
            FROM agenda_items ai
            JOIN persons p ON p.id = ai.person_id
            WHERE ai.agenda_id = ?
            ORDER BY datetime(ai.created_at) DESC
            """,
            (agenda_id,),
        ).fetchall()
        return render_template(
            "agenda_form.html",
            agenda=agenda,
            items=items,
            assistance_types=AGENDA_ASSISTANCE_TYPES,
        )

    @app.route("/agenda/<int:agenda_id>/update", methods=["POST"])
    def update_agenda(agenda_id):
        fetch_agenda(agenda_id)
        title = request.form.get("title", "").strip()
        meeting_date = request.form.get("meeting_date", "").strip() or None
        description = None

        if not title:
            flash("Gündem başlığı zorunludur.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=agenda_id))

        db = get_db()
        now = datetime.utcnow().isoformat()
        db.execute(
            """
            UPDATE agendas
            SET title = ?, meeting_date = ?, description = ?, updated_at = ?
            WHERE id = ?
            """,
            (title, meeting_date, description, now, agenda_id),
        )
        db.commit()
        flash("Gündem güncellendi.", "success")
        return redirect(url_for("edit_agenda", agenda_id=agenda_id))

    @app.route("/agenda/<int:agenda_id>/delete", methods=["POST"])
    def delete_agenda(agenda_id):
        fetch_agenda(agenda_id)
        db = get_db()
        db.execute("DELETE FROM agendas WHERE id = ?", (agenda_id,))
        db.commit()
        flash("Gündem silindi.", "info")
        return redirect(url_for("list_agendas"))

    @app.route("/agenda/<int:agenda_id>/items", methods=["POST"])
    def add_agenda_item(agenda_id):
        fetch_agenda(agenda_id)
        db = get_db()
        file_no = request.form.get("file_no", "").strip()
        application_dates = [value.strip() for value in request.form.getlist("application_date[]")]
        assistance_types = [value.strip() for value in request.form.getlist("assistance_type[]")]
        notes_list = [value.strip() for value in request.form.getlist("notes[]")]

        if not application_dates and "application_date" in request.form:
            application_dates = [request.form.get("application_date", "").strip()]
        if not assistance_types and "assistance_type" in request.form:
            assistance_types = [(request.form.get("assistance_type") or "").strip()]
        if not notes_list and "notes" in request.form:
            notes_list = [request.form.get("notes", "").strip()]

        if not file_no:
            flash("Dosya numarası zorunludur.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=agenda_id))

        person = db.execute(
            "SELECT id, full_name FROM persons WHERE file_no = ?", (file_no,)
        ).fetchone()

        if person is None:
            flash("Dosya numarası bulunamadı.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=agenda_id))

        max_len = max(len(application_dates), len(assistance_types), len(notes_list))
        entries = []
        for idx in range(max_len):
            raw_date = application_dates[idx] if idx < len(application_dates) else ""
            raw_type = assistance_types[idx] if idx < len(assistance_types) else ""
            raw_note = notes_list[idx] if idx < len(notes_list) else ""

            entry = {
                "application_date": raw_date or None,
                "assistance_type": raw_type,
                "notes": raw_note,
            }

            if not any(entry.values()):
                continue

            if entry["assistance_type"] and entry["assistance_type"] not in AGENDA_ASSISTANCE_TYPES:
                flash("Geçersiz yardım türü seçildi.", "danger")
                return redirect(url_for("edit_agenda", agenda_id=agenda_id))

            entries.append(entry)

        if not entries:
            flash("En az bir başvuru satırı doldurulmalıdır.", "warning")
            return redirect(url_for("edit_agenda", agenda_id=agenda_id))

        for entry in entries:
            now = datetime.utcnow().isoformat()
            db.execute(
                """
                INSERT INTO agenda_items (
                    agenda_id, person_id, application_date, assistance_type, notes, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    agenda_id,
                    person["id"],
                    entry["application_date"],
                    entry["assistance_type"],
                    entry["notes"],
                    now,
                ),
            )
        db.commit()
        if len(entries) > 1:
            flash(f"Gündeme {len(entries)} başvuru kaydı eklendi.", "success")
        else:
            flash("Gündeme kayıt eklendi.", "success")
        return redirect(url_for("edit_agenda", agenda_id=agenda_id))

    @app.route("/agenda_item/<int:item_id>/update", methods=["POST"])
    def update_agenda_item(item_id):
        db = get_db()
        item = db.execute(
            "SELECT id, agenda_id FROM agenda_items WHERE id = ?", (item_id,)
        ).fetchone()
        if item is None:
            abort(404, description="Gündem kaydı bulunamadı.")

        file_no = request.form.get("file_no", "").strip()
        assistance_type = request.form.get("assistance_type") or ""
        application_date = request.form.get("application_date", "").strip() or None
        notes = request.form.get("notes", "").strip()

        if not file_no:
            flash("Dosya numarası zorunludur.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=item["agenda_id"]))

        person = db.execute(
            "SELECT id FROM persons WHERE file_no = ?", (file_no,)
        ).fetchone()

        if person is None:
            flash("Dosya numarası bulunamadı.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=item["agenda_id"]))

        if assistance_type and assistance_type not in AGENDA_ASSISTANCE_TYPES:
            flash("Geçersiz yardım türü seçildi.", "danger")
            return redirect(url_for("edit_agenda", agenda_id=item["agenda_id"]))

        db.execute(
            """
            UPDATE agenda_items
            SET person_id = ?, application_date = ?, assistance_type = ?, notes = ?
            WHERE id = ?
            """,
            (person["id"], application_date, assistance_type, notes, item_id),
        )
        db.commit()
        flash("Gündem kaydı güncellendi.", "success")
        return redirect(url_for("edit_agenda", agenda_id=item["agenda_id"]))

    @app.route("/agenda_item/<int:item_id>/delete", methods=["POST"])
    def delete_agenda_item(item_id):
        db = get_db()
        item = db.execute(
            "SELECT agenda_id FROM agenda_items WHERE id = ?", (item_id,)
        ).fetchone()
        if item is None:
            abort(404, description="Gündem kaydı bulunamadı.")
        db.execute("DELETE FROM agenda_items WHERE id = ?", (item_id,))
        db.commit()
        flash("Gündem kaydı silindi.", "info")
        return redirect(url_for("edit_agenda", agenda_id=item["agenda_id"]))

    def build_agenda_report_rows(agenda_id):
        db = get_db()
        rows = db.execute(
            """
            SELECT
                ai.id AS agenda_item_id,
                ai.application_date,
                ai.assistance_type,
                ai.notes,
                ai.created_at AS created_at,
                p.file_no,
                p.full_name,
                p.national_id,
                p.phone,
                p.address,
                p.social_security,
                p.household_size,
                p.children_count,
                p.student_count,
                p.household_income,
                p.per_capita_income
            FROM agenda_items ai
            JOIN persons p ON p.id = ai.person_id
            WHERE ai.agenda_id = ?
            ORDER BY datetime(ai.created_at)
            """,
            (agenda_id,),
        ).fetchall()

        report_rows = []
        for row in rows:
            report_rows.append(
                {
                    "agenda_item_id": row["agenda_item_id"],
                    "application_date": row["application_date"],
                    "assistance_type": row["assistance_type"],
                    "notes": (row["notes"] or "").strip(),
                    "created_at": row["created_at"],
                    "file_no": row["file_no"],
                    "full_name": row["full_name"],
                    "national_id": row["national_id"],
                    "phone": row["phone"],
                    "address": row["address"],
                    "social_security": row["social_security"],
                    "household_size": row["household_size"],
                    "children_count": row["children_count"],
                    "student_count": row["student_count"],
                    "household_income": row["household_income"],
                    "per_capita_income": row["per_capita_income"],
                }
            )
        return report_rows

    @app.route("/agenda/<int:agenda_id>/report")
    def agenda_report(agenda_id):
        agenda = fetch_agenda(agenda_id)
        report_rows = build_agenda_report_rows(agenda_id)
        return render_template(
            "agenda_report.html",
            agenda=agenda,
            report_rows=report_rows,
        )

    @app.route("/agenda/<int:agenda_id>/report.xlsx")
    def agenda_report_excel(agenda_id):
        agenda = fetch_agenda(agenda_id)
        report_rows = build_agenda_report_rows(agenda_id)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Gundem Raporu"

        headers = [
            "Dosya No",
            "Ad Soyad",
            "Basvuru Tarihi",
            "Yardim Turu",
            "Notlar",
            "Adres",
            "Sosyal Guvence",
            "Hane Nufusu",
            "Cocuk Sayisi",
            "Ogrenci Sayisi",
            "Hane Geliri",
            "Kisi Basina Gelir",
        ]
        worksheet.append(headers)

        column_widths = [len(header) for header in headers]
        for row in report_rows:
            values = [
                row["file_no"] or "",
                row["full_name"] or "",
                row["application_date"] or "",
                row["assistance_type"] or "",
                row["notes"] or "",
                row["address"] or "",
                row["social_security"] or "",
                row["household_size"] if row["household_size"] is not None else "",
                row["children_count"] if row["children_count"] is not None else "",
                row["student_count"] if row["student_count"] is not None else "",
                row["household_income"] if row["household_income"] is not None else "",
                row["per_capita_income"] if row["per_capita_income"] is not None else "",
            ]
            worksheet.append(values)
            for index, value in enumerate(values):
                width = len(str(value)) if value is not None else 0
                if width > column_widths[index]:
                    column_widths[index] = width

        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font

        for index, width in enumerate(column_widths, start=1):
            adjusted = min(max(width + 2, 12), 45)
            worksheet.column_dimensions[get_column_letter(index)].width = adjusted

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename_base = secure_filename(agenda["title"] or f"gundem_{agenda_id}")
        if not filename_base:
            filename_base = f"gundem_{agenda_id}"
        filename = f"{filename_base}_rapor.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/agenda/<int:agenda_id>/presentation")
    def agenda_presentation(agenda_id):
        agenda = fetch_agenda(agenda_id)
        db = get_db()
        items = db.execute(
            """
            SELECT
                ai.id AS agenda_item_id,
                ai.agenda_id,
                ai.application_date,
                ai.assistance_type,
                ai.notes,
                ai.created_at AS agenda_item_created_at,
                p.id AS person_id,
                p.file_no,
                p.full_name,
                p.national_id,
                p.birth_date,
                p.spouse_name,
                p.household_size,
                p.children_count,
                p.student_count,
                p.phone,
                p.address,
                p.social_security,
                p.disability_status,
                p.disability_rate,
                p.central_assistance,
                p.household_description,
                p.household_income,
                p.per_capita_income,
                p.profile_photo IS NOT NULL AS has_profile_photo
            FROM agenda_items ai
            JOIN persons p ON p.id = ai.person_id
            WHERE ai.agenda_id = ?
            ORDER BY datetime(ai.created_at)
            """,
            (agenda_id,),
        ).fetchall()

        slides = []
        for row in items:
            person_id = row["person_id"]
            assistance_records = db.execute(
                """
                SELECT assistance_type, assistance_date, assistance_amount
                FROM assistance_records
                WHERE person_id = ?
                ORDER BY datetime(assistance_date) DESC, id DESC
                LIMIT 2
                """,
                (person_id,),
            ).fetchall()
            household_images = db.execute(
                """
                SELECT id, filename
                FROM household_images
                WHERE person_id = ?
                ORDER BY id
                """,
                (person_id,),
            ).fetchall()
            central_assistance = []
            if row["central_assistance"]:
                try:
                    central_assistance = json.loads(row["central_assistance"])
                except json.JSONDecodeError:
                    central_assistance = []

            slides.append(
                {
                    "agenda_item_id": row["agenda_item_id"],
                    "application_date": row["application_date"],
                    "assistance_type": row["assistance_type"],
                    "notes": row["notes"],
                    "person": {
                        "id": person_id,
                        "file_no": row["file_no"],
                        "full_name": row["full_name"],
                        "national_id": row["national_id"],
                        "birth_date": row["birth_date"],
                        "age": calculate_age(row["birth_date"]),
                        "spouse_name": row["spouse_name"],
                        "household_size": row["household_size"],
                        "children_count": row["children_count"],
                        "student_count": row["student_count"],
                        "phone": row["phone"],
                        "address": row["address"],
                        "social_security": row["social_security"],
                        "disability_status": row["disability_status"],
                        "disability_rate": row["disability_rate"],
                        "household_description": (row["household_description"] or "").strip(),
                        "household_description_lines": normalize_description(row["household_description"]),
                        "household_income": row["household_income"],
                        "per_capita_income": row["per_capita_income"],
                        "has_profile_photo": bool(row["has_profile_photo"]),
                    },
                    "central_assistance": central_assistance,
                    "assistance_records": assistance_records,
                    "household_images": household_images,
                    "agenda_item_created_at": row["agenda_item_created_at"],
                }
            )

        return render_template(
            "agenda_presentation.html",
            agenda=agenda,
            slides=slides,
        )

    @app.route("/person/new", methods=["GET", "POST"])
    def create_person():
        if request.method == "POST":
            return handle_person_form()

        return render_template(
            "person_form.html",
            person=None,
            assistance_records=[],
            household_images=[],
            social_security_options=SOCIAL_SECURITY_OPTIONS,
            disability_options=DISABILITY_OPTIONS,
            assistance_types=ASSISTANCE_TYPES,
            central_assistance_options=CENTRAL_ASSISTANCE_OPTIONS,
        )

    @app.route("/person/<int:person_id>/edit", methods=["GET", "POST"])
    def edit_person(person_id):
        person = fetch_person(person_id)
        db = get_db()
        assistance_records = db.execute(
            "SELECT * FROM assistance_records WHERE person_id = ? ORDER BY id",
            (person_id,),
        ).fetchall()
        household_images = db.execute(
            "SELECT id, filename FROM household_images WHERE person_id = ? ORDER BY id",
            (person_id,),
        ).fetchall()

        if request.method == "POST":
            return handle_person_form(person_id=person_id)

        selected_central = (
            json.loads(person["central_assistance"])
            if person["central_assistance"]
            else []
        )
        return render_template(
            "person_form.html",
            person=person,
            assistance_records=assistance_records,
            household_images=household_images,
            selected_central=selected_central,
            social_security_options=SOCIAL_SECURITY_OPTIONS,
            disability_options=DISABILITY_OPTIONS,
            assistance_types=ASSISTANCE_TYPES,
            central_assistance_options=CENTRAL_ASSISTANCE_OPTIONS,
        )

    def handle_person_form(person_id=None):
        db = get_db()
        form = request.form
        file_no = form.get("file_no", "").strip()
        full_name = form.get("full_name", "").strip()
        national_id = form.get("national_id", "").strip()
        birth_date = form.get("birth_date", "").strip() or None
        spouse_name = form.get("spouse_name", "").strip()
        household_size = parse_int(form.get("household_size"))
        children_count = parse_int(form.get("children_count"))
        student_count = parse_int(form.get("student_count"))
        phone = form.get("phone", "").strip()
        address = form.get("address", "").strip()
        social_security = form.get("social_security") or None
        disability_status = form.get("disability_status") or None
        disability_rate = (
            form.get("disability_rate", "").strip()
            if disability_status == "Var"
            else None
        )
        household_description = form.get("household_description", "").strip()
        household_income = parse_float(form.get("household_income"))
        per_capita_income = parse_float(form.get("per_capita_income"))
        central_assistance = form.getlist("central_assistance")
        central_assistance_json = json.dumps(central_assistance, ensure_ascii=False)

        if not file_no or not full_name:
            flash("Dosya numarası ve Adı Soyadı alanları zorunludur.", "danger")
            return redirect(request.url)

        profile_file = request.files.get("profile_photo")
        profile_data = None
        profile_filename = None
        profile_mimetype = None

        if profile_file and profile_file.filename:
            profile_filename = secure_filename(profile_file.filename)
            profile_data = profile_file.read()
            profile_mimetype = profile_file.mimetype

        assistance_types = form.getlist("assistance_type[]")
        assistance_dates = form.getlist("assistance_date[]")
        assistance_amounts = form.getlist("assistance_amount[]")

        now = datetime.utcnow().isoformat()

        try:
            if person_id is None:
                cursor = db.execute(
                    """
                    INSERT INTO persons (
                        file_no, full_name, national_id, birth_date, spouse_name, household_size,
                        children_count, student_count, phone, address, social_security,
                        disability_status, disability_rate, central_assistance,
                        household_description, household_income, per_capita_income,
                        profile_photo, profile_photo_filename,
                        profile_photo_mimetype, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        file_no,
                        full_name,
                        national_id,
                        birth_date,
                        spouse_name,
                        household_size,
                        children_count,
                        student_count,
                        phone,
                        address,
                        social_security,
                        disability_status,
                        disability_rate,
                        central_assistance_json,
                        household_description,
                        household_income,
                        per_capita_income,
                        profile_data,
                        profile_filename,
                        profile_mimetype,
                        now,
                        now,
                    ),
                )
                person_id = cursor.lastrowid
            else:
                params = [
                    file_no,
                    full_name,
                    national_id,
                    birth_date,
                    spouse_name,
                    household_size,
                    children_count,
                    student_count,
                    phone,
                    address,
                    social_security,
                    disability_status,
                    disability_rate,
                    central_assistance_json,
                    household_description,
                    household_income,
                    per_capita_income,
                    now,
                    person_id,
                ]
                db.execute(
                    """
                    UPDATE persons
                        SET file_no = ?, full_name = ?, national_id = ?, birth_date = ?, spouse_name = ?,
                        household_size = ?, children_count = ?, student_count = ?,
                        phone = ?, address = ?, social_security = ?, disability_status = ?,
                        disability_rate = ?, central_assistance = ?, household_description = ?,
                        household_income = ?, per_capita_income = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    params,
                )

                if profile_data is not None:
                    db.execute(
                        """
                        UPDATE persons
                        SET profile_photo = ?, profile_photo_filename = ?, profile_photo_mimetype = ?
                        WHERE id = ?
                        """,
                        (profile_data, profile_filename, profile_mimetype, person_id),
                    )

                delete_image_ids = form.getlist("delete_image_ids")
                if delete_image_ids:
                    placeholders = ",".join("?" for _ in delete_image_ids)
                    db.execute(
                        f"DELETE FROM household_images WHERE person_id = ? AND id IN ({placeholders})",
                        [person_id, *delete_image_ids],
                    )

                db.execute(
                    "DELETE FROM assistance_records WHERE person_id = ?", (person_id,)
                )

            for a_type, a_date, a_amount in zip(
                assistance_types, assistance_dates, assistance_amounts
            ):
                a_type = (a_type or "").strip()
                a_date = (a_date or "").strip()
                a_amount = (a_amount or "").strip()
                if not a_type and not a_date and not a_amount:
                    continue
                amount_value = parse_float(a_amount)
                db.execute(
                    """
                    INSERT INTO assistance_records (
                        person_id, assistance_type, assistance_date, assistance_amount
                    )
                    VALUES (?, ?, ?, ?)
                    """,
                    (person_id, a_type, a_date, amount_value),
                )

            household_files = request.files.getlist("household_images")
            for file in household_files:
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    data = file.read()
                    db.execute(
                        """
                        INSERT INTO household_images (
                            person_id, image_data, filename, mimetype, created_at
                        )
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (person_id, data, filename, file.mimetype, now),
                    )

            db.commit()
        except sqlite3.IntegrityError:
            flash("Bu dosya numarası zaten kayıtlı.", "danger")
            return redirect(request.url)

        flash("Kayıt başarıyla kaydedildi.", "success")
        return redirect(url_for("index"))

    @app.route("/person/<int:person_id>/delete", methods=["POST"])
    def delete_person(person_id):
        db = get_db()
        db.execute("DELETE FROM persons WHERE id = ?", (person_id,))
        db.commit()
        flash("Kayıt silindi.", "info")
        return redirect(url_for("index"))

    @app.route("/person/<int:person_id>/profile_photo")
    def profile_photo(person_id):
        person = fetch_person(person_id)
        if not person["profile_photo"]:
            abort(404)
        return send_file(
            BytesIO(person["profile_photo"]),
            mimetype=person["profile_photo_mimetype"] or "application/octet-stream",
            download_name=person["profile_photo_filename"] or "profil.jpg",
        )

    @app.route("/household_image/<int:image_id>")
    def household_image(image_id):
        db = get_db()
        image = db.execute(
            "SELECT image_data, filename, mimetype FROM household_images WHERE id = ?",
            (image_id,),
        ).fetchone()
        if image is None:
            abort(404)
        return send_file(
            BytesIO(image["image_data"]),
            mimetype=image["mimetype"] or "application/octet-stream",
            download_name=image["filename"] or "hane.jpg",
        )

    @app.route("/check_file_no")
    def check_file_no():
        file_no = request.args.get("file_no", "").strip()
        if not file_no:
            return jsonify({"exists": False})
        db = get_db()
        exists = (
            db.execute(
                "SELECT 1 FROM persons WHERE file_no = ? LIMIT 1", (file_no,)
            ).fetchone()
            is not None
        )
        return jsonify({"exists": exists})

    @app.route("/lookup_person")
    def lookup_person():
        file_no = request.args.get("file_no", "").strip()
        if not file_no:
            return jsonify({"found": False})
        db = get_db()
        person = db.execute(
            """
            SELECT id, file_no, full_name, phone, address, social_security
            FROM persons
            WHERE file_no = ?
            """,
            (file_no,),
        ).fetchone()
        if person is None:
            return jsonify({"found": False})
        return jsonify(
            {
                "found": True,
                "person": {
                    "id": person["id"],
                    "file_no": person["file_no"],
                    "full_name": person["full_name"],
                    "phone": person["phone"],
                    "address": person["address"],
                    "social_security": person["social_security"],
                },
            }
        )

    @app.template_filter("datetimeformat")
    def datetime_format_filter(value):
        if not value:
            return "—"
        try:
            dt = datetime.fromisoformat(value)
        except ValueError:
            return value
        return dt.strftime("%d.%m.%Y %H:%M")

    @app.template_filter("dateformat")
    def date_format_filter(value):
        if not value:
            return "—"
        try:
            dt = datetime.fromisoformat(value)
        except ValueError:
            return value
        return dt.strftime("%d.%m.%Y")

    @app.template_filter("currency")
    def currency_format_filter(value):
        if value is None or value == "":
            return "—"
        try:
            number = float(value)
        except (TypeError, ValueError):
            return value
        formatted = f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{formatted} TL"

    return app


def parse_int(value):
    try:
        return int(value) if value is not None and value != "" else None
    except ValueError:
        return None


def parse_float(value):
    try:
        return float(value.replace(",", ".")) if value else None
    except (ValueError, AttributeError):
        return None


BULLET_PREFIX_RE = re.compile(r"^[\-\*\u2022\u2023\u25AA\u25CF\s]+")
STAR_BREAK_RE = re.compile(r"(?<!^)(?<!\n)(\*+)")
NUMBER_BREAK_RE = re.compile(r"(?<!^)(?<!\n)(\d{1,2}\s*[.\-):])")
UPPER_BREAK_RE = re.compile(r"(?<!^)(?<!\n)\s{3,}(?=[A-ZÇĞİÖŞÜ])")
CURRENCY_BREAK_RE = re.compile(r"(?<=TL)\s+(?=[A-ZÇĞİÖŞÜ])")


def normalize_description(text):
    if not text:
        return []
    normalized = (text or "").replace("\r\n", "\n")
    normalized = STAR_BREAK_RE.sub(r"\n\1", normalized)
    normalized = NUMBER_BREAK_RE.sub(r"\n\1", normalized)
    normalized = UPPER_BREAK_RE.sub("\n", normalized)
    normalized = CURRENCY_BREAK_RE.sub("\n", normalized)
    normalized = re.sub(r"\n{2,}", "\n", normalized)
    segments = []
    for raw in re.split(r"\n+", normalized):
        stripped = raw.strip()
        if not stripped:
            continue
        original = stripped
        stripped = BULLET_PREFIX_RE.sub("", stripped)
        if original.startswith("**") and not stripped.startswith("**"):
            stripped = "**" + stripped
        if stripped:
            segments.append(stripped)
    if len(segments) <= 1:
        base_text = segments[0] if segments else normalized.strip()
        if base_text:
            base_text = STAR_BREAK_RE.sub(r"\n\1", base_text)
            base_text = NUMBER_BREAK_RE.sub(r"\n\1", base_text)
            base_text = UPPER_BREAK_RE.sub("\n", base_text)
            base_text = CURRENCY_BREAK_RE.sub("\n", base_text)
            base_text = re.sub(r"\n{2,}", "\n", base_text)
            candidates = re.split(r"(?<=[.!?;:])\s+", base_text)
            refined = []
            for candidate in candidates:
                original = candidate.strip()
                cleaned = BULLET_PREFIX_RE.sub("", original)
                if original.startswith("**") and not cleaned.startswith("**"):
                    cleaned = "**" + cleaned
                if not cleaned:
                    continue
                if len(cleaned) > 110 and "," in cleaned:
                    refined.extend(
                        part.strip()
                        for part in cleaned.split(",")
                        if part.strip()
                    )
                else:
                    refined.append(cleaned)
            segments = refined
    return segments


def calculate_age(birth_date_str):
    if not birth_date_str:
        return None
    try:
        birth_date = datetime.fromisoformat(birth_date_str).date()
    except ValueError:
        try:
            birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
        except ValueError:
            return None
    today = datetime.utcnow().date()
    age = today.year - birth_date.year - (
        (today.month, today.day) < (birth_date.month, birth_date.day)
    )
    return age if age >= 0 else None


app = create_app()


if __name__ == "__main__":
    run_host = os.environ.get("FLASK_RUN_HOST", "0.0.0.0")
    run_port = os.environ.get("FLASK_RUN_PORT", "5000")
    try:
        port_value = int(run_port)
    except ValueError:
        port_value = 5000
    debug_flag = os.environ.get("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host=run_host, port=port_value, debug=debug_flag)
